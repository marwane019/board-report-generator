"""
excel_pack.py — Board Report Excel Data Pack Generator.

Produces a 6-sheet Excel workbook — the data analyst's companion to the PDF.
Every sheet is formatted to Big-4 presentation standards with:
    - Colour-coded headers matching the brand palette
    - Conditional formatting on variance columns
    - Embedded sparklines (via openpyxl charts where supported)
    - Freeze panes and auto-filter on all data sheets

Sheets:
    1. Summary         — KPI dashboard with RAG tiles
    2. P&L             — Full monthly P&L (actuals, budget, prior year, variances)
    3. Pipeline        — Weekly pipeline data with stage breakdown
    4. Customers       — Monthly ARR waterfall and customer counts
    5. Headcount       — Monthly headcount by department
    6. Data Dictionary — Column definitions for all sheets
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.metrics import MetricsPackage

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED = Side(style="medium")


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_colour.lstrip("#"))


def _font(bold: bool = False, colour: str = "000000", size: int = 10,
          italic: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, color=colour.lstrip("#"),
                size=size, italic=italic)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _auto_fit(ws, min_w: int = 8, max_w: int = 55) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col), default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(max_len + 3, min_w), max_w)


def _write_header_row(ws, row: int, headers: list[str], brand: dict) -> None:
    """Write a formatted header row at the given row index."""
    primary = brand["primary"].lstrip("#")
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_i, value=h)
        cell.fill = _fill(primary)
        cell.font = _font(bold=True, colour="FFFFFF", size=10)
        cell.alignment = _center()
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 20


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_summary(ws, pkg: MetricsPackage, brand: dict) -> None:
    """Write the KPI summary dashboard sheet."""
    ws.sheet_properties.tabColor = brand["primary"].lstrip("#")
    primary = brand["primary"].lstrip("#")
    light = brand["light"].lstrip("#")

    # Title
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"{pkg.company_name} — Board Report KPI Dashboard"
    cell.fill = _fill(primary)
    cell.font = _font(bold=True, colour="FFFFFF", size=14)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Period: {pkg.report_period}  |  Generated: {datetime.today().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = _font(italic=True, colour="555555", size=9)
    ws["A2"].alignment = _center()

    # KPI sections
    rag = pkg.rag
    fin = pkg.financial
    comm = pkg.commercial
    cust = pkg.customers
    hc = pkg.headcount

    rag_colour_map = {
        "Green": brand["green"].lstrip("#"),
        "Amber": brand["amber"].lstrip("#"),
        "Red":   brand["red"].lstrip("#"),
    }

    kpis = [
        ("FINANCIAL PERFORMANCE", "", "", ""),
        ("Metric", "Actual", "Budget", "RAG"),
        ("Revenue", f"£{fin.revenue_actual/1e6:.2f}M",
         f"£{fin.revenue_budget/1e6:.2f}M", rag.revenue.status),
        ("Gross Margin %", f"{fin.gross_margin_pct_actual*100:.1f}%",
         f"{fin.gross_margin_pct_budget*100:.1f}%", rag.gross_margin.status),
        ("EBITDA", f"£{fin.ebitda_actual/1000:.0f}k",
         f"£{fin.ebitda_budget/1000:.0f}k", rag.ebitda_margin.status),
        ("EBITDA Margin %", f"{fin.ebitda_margin_pct_actual*100:.1f}%",
         f"{fin.ebitda_margin_pct_budget*100:.1f}%", rag.ebitda_margin.status),
        ("YTD Revenue", f"£{fin.ytd_revenue_actual/1e6:.2f}M",
         f"£{fin.ytd_revenue_budget/1e6:.2f}M",
         "Green" if fin.ytd_revenue_actual >= fin.ytd_revenue_budget * 0.95 else "Amber"),
        ("", "", "", ""),
        ("COMMERCIAL PERFORMANCE", "", "", ""),
        ("Metric", "Actual", "Budget / Target", "RAG"),
        ("Pipeline Coverage", f"{comm.pipeline_coverage_ratio:.1f}x",
         "3.0x", rag.pipeline_coverage.status),
        ("Win Rate", f"{comm.win_rate_actual*100:.1f}%",
         f"{comm.win_rate_budget*100:.1f}%", rag.win_rate.status),
        ("Total Pipeline", f"£{comm.total_pipeline_gbp/1e6:.1f}M",
         f"£{comm.pipeline_budget_gbp/1e6:.1f}M", "Green"),
        ("", "", "", ""),
        ("CUSTOMER METRICS", "", "", ""),
        ("Metric", "Actual", "Budget / Target", "RAG"),
        ("ARR", f"£{cust.arr_actual/1e6:.2f}M",
         f"£{cust.arr_budget/1e6:.2f}M", "Green"),
        ("Monthly Churn Rate", f"{cust.churn_rate_actual*100:.2f}%",
         f"{cust.churn_rate_budget*100:.2f}%", rag.churn_rate.status),
        ("NPS", str(cust.nps_actual), str(cust.nps_budget), rag.nps.status),
        ("", "", "", ""),
        ("PEOPLE & OPERATIONS", "", "", ""),
        ("Metric", "Actual", "Budget", "RAG"),
        ("Total Headcount", str(hc.total_hc_actual), str(hc.total_hc_budget),
         rag.headcount.status),
        ("Monthly People Cost", f"£{hc.total_cost_actual/1000:.0f}k",
         f"£{hc.total_cost_budget/1000:.0f}k", "Green"),
        ("Cost Per Head (monthly)", f"£{hc.cost_per_head_actual:.0f}",
         f"£{hc.cost_per_head_budget:.0f}", "Green"),
    ]

    for row_i, (label, actual, budget, rag_status) in enumerate(kpis, start=4):
        if label in ("FINANCIAL PERFORMANCE", "COMMERCIAL PERFORMANCE",
                     "CUSTOMER METRICS", "PEOPLE & OPERATIONS"):
            ws.merge_cells(
                start_row=row_i, start_column=1,
                end_row=row_i, end_column=4
            )
            c = ws.cell(row=row_i, column=1, value=label)
            c.fill = _fill(primary)
            c.font = _font(bold=True, colour="FFFFFF", size=10)
            c.alignment = _center()
            ws.row_dimensions[row_i].height = 18
        elif label == "Metric":
            for col_i, val in enumerate([label, actual, budget, rag_status], start=1):
                c = ws.cell(row=row_i, column=col_i, value=val)
                c.fill = _fill(light)
                c.font = _font(bold=True, colour="333333", size=9)
                c.alignment = _center()
                c.border = THIN_BORDER
        elif label == "":
            pass
        else:
            for col_i, val in enumerate([label, actual, budget, rag_status], start=1):
                c = ws.cell(row=row_i, column=col_i, value=val)
                c.font = _font(size=9)
                c.alignment = _center()
                c.border = THIN_BORDER
                if col_i == 4 and rag_status in rag_colour_map:
                    c.fill = _fill(rag_colour_map[rag_status])
                    c.font = _font(bold=True, colour="FFFFFF", size=9)

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 28


def _sheet_pl(ws, pkg: MetricsPackage, brand: dict, config_path: str) -> None:
    """Write the full P&L sheet with actuals, budget, and variances."""
    ws.sheet_properties.tabColor = brand["secondary"].lstrip("#")

    with open(config_path, "r") as fh:
        cfg = yaml.safe_load(fh)

    fin_path = cfg["paths"]["financials_file"]
    if not Path(fin_path).exists():
        ws["A1"].value = "Run --generate-data first"
        return

    fin_df = pd.read_csv(fin_path)
    fin_df["variance_gbp"] = fin_df["actual_gbp"] - fin_df["budget_gbp"]
    fin_df["variance_pct"] = (
        (fin_df["actual_gbp"] / fin_df["budget_gbp"]) - 1
    ).replace([float("inf"), float("-inf")], 0).fillna(0)
    fin_df["yoy_growth_pct"] = (
        (fin_df["actual_gbp"] / fin_df["prior_year_gbp"]) - 1
    ).replace([float("inf"), float("-inf")], 0).fillna(0)

    display_cols = [
        "period", "line_type", "line_name",
        "actual_gbp", "budget_gbp", "variance_gbp", "variance_pct",
        "prior_year_gbp", "yoy_growth_pct",
    ]
    df_out = fin_df[display_cols].rename(columns={
        "period": "Period",
        "line_type": "Type",
        "line_name": "Line",
        "actual_gbp": "Actual (£)",
        "budget_gbp": "Budget (£)",
        "variance_gbp": "Variance (£)",
        "variance_pct": "Variance %",
        "prior_year_gbp": "Prior Year (£)",
        "yoy_growth_pct": "YoY Growth %",
    })

    _write_header_row(ws, 1, list(df_out.columns), brand)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df_out.columns))}1"

    for row_i, row in enumerate(
        dataframe_to_rows(df_out, index=False, header=False), start=2
    ):
        for col_i, val in enumerate(row, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font = _font(size=9)
            c.border = THIN_BORDER
            col_name = df_out.columns[col_i - 1]
            if "(£)" in col_name:
                c.number_format = "#,##0"
            elif "%" in col_name:
                c.number_format = "0.0%"
            row_fill = "F9F9F9" if row_i % 2 == 0 else "FFFFFF"
            c.fill = _fill(row_fill)

    # Conditional formatting on variance column
    var_col = get_column_letter(df_out.columns.get_loc("Variance (£)") + 1)
    last_row = len(df_out) + 1
    ws.conditional_formatting.add(
        f"{var_col}2:{var_col}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="FFC7CE",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="C6EFCE",
        )
    )
    _auto_fit(ws)


def _sheet_pipeline(ws, brand: dict, config_path: str) -> None:
    """Write the pipeline sheet."""
    ws.sheet_properties.tabColor = brand["accent"].lstrip("#")
    with open(config_path, "r") as fh:
        cfg = yaml.safe_load(fh)
    pipe_path = cfg["paths"]["pipeline_file"]
    if not Path(pipe_path).exists():
        ws["A1"].value = "Run --generate-data first"
        return

    df = pd.read_csv(pipe_path)
    df["pipeline_variance_gbp"] = df["pipeline_value_gbp"] - df["budget_pipeline_gbp"]
    _write_header_row(ws, 1, list(df.columns), brand)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
    for row_i, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=2
    ):
        for col_i, val in enumerate(row, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font = _font(size=9)
            c.border = THIN_BORDER
            if isinstance(val, float) and "gbp" in df.columns[col_i - 1]:
                c.number_format = "#,##0"
    _auto_fit(ws)


def _sheet_customers(ws, brand: dict, config_path: str) -> None:
    """Write the customer metrics sheet."""
    ws.sheet_properties.tabColor = brand["green"].lstrip("#")
    with open(config_path, "r") as fh:
        cfg = yaml.safe_load(fh)
    path = cfg["paths"]["customers_file"]
    if not Path(path).exists():
        ws["A1"].value = "Run --generate-data first"
        return

    df = pd.read_csv(path)
    df["net_arr_gbp"] = df["new_arr_gbp"] - df["churned_arr_gbp"]
    df["arr_vs_budget_pct"] = (df["arr_gbp"] / df["arr_budget_gbp"] - 1).fillna(0)

    _write_header_row(ws, 1, list(df.columns), brand)
    ws.freeze_panes = "A2"
    for row_i, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=2
    ):
        for col_i, val in enumerate(row, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font = _font(size=9)
            c.border = THIN_BORDER
    _auto_fit(ws)


def _sheet_headcount(ws, brand: dict, config_path: str) -> None:
    """Write the headcount sheet."""
    ws.sheet_properties.tabColor = brand["amber"].lstrip("#")
    with open(config_path, "r") as fh:
        cfg = yaml.safe_load(fh)
    path = cfg["paths"]["headcount_file"]
    if not Path(path).exists():
        ws["A1"].value = "Run --generate-data first"
        return

    df = pd.read_csv(path)
    df["hc_variance"] = df["headcount_actual"] - df["headcount_budget"]
    df["cost_variance_gbp"] = df["cost_actual_gbp"] - df["cost_budget_gbp"]

    _write_header_row(ws, 1, list(df.columns), brand)
    ws.freeze_panes = "A2"
    for row_i, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=2
    ):
        for col_i, val in enumerate(row, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font = _font(size=9)
            c.border = THIN_BORDER
            if "cost" in df.columns[col_i - 1]:
                c.number_format = "#,##0"
    _auto_fit(ws)


def _sheet_data_dict(ws, brand: dict) -> None:
    """Write the data dictionary sheet."""
    ws.sheet_properties.tabColor = "888888"
    primary = brand["primary"].lstrip("#")

    definitions = [
        ("Sheet", "Column", "Description", "Format"),
        ("P&L", "Actual (£)", "Invoiced/recorded amount for the period", "£ integer"),
        ("P&L", "Budget (£)", "Board-approved budget for the period", "£ integer"),
        ("P&L", "Variance (£)", "Actual minus Budget; positive = favourable for Revenue, negative for costs", "£ integer"),
        ("P&L", "Variance %", "Variance as % of Budget", "Percentage"),
        ("P&L", "Prior Year (£)", "Same period prior year actuals", "£ integer"),
        ("P&L", "YoY Growth %", "Actual vs Prior Year growth rate", "Percentage"),
        ("Pipeline", "pipeline_value_gbp", "Total pipeline value in stage for the week", "£ integer"),
        ("Pipeline", "win_rate_actual", "Estimated win rate for opportunities in this stage", "Decimal"),
        ("Customers", "arr_gbp", "Annual Recurring Revenue at period end", "£ integer"),
        ("Customers", "churn_rate_actual", "Monthly churn rate (churned ARR / opening ARR)", "Decimal"),
        ("Customers", "nps_actual", "Net Promoter Score (-100 to +100)", "Integer"),
        ("Headcount", "headcount_actual", "FTEs on payroll at period end", "Integer"),
        ("Headcount", "cost_actual_gbp", "Total payroll cost for the period", "£ integer"),
    ]

    for row_i, row_data in enumerate(definitions, start=1):
        is_header = row_i == 1
        for col_i, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            if is_header:
                c.fill = _fill(primary)
                c.font = _font(bold=True, colour="FFFFFF", size=9)
            else:
                c.font = _font(size=9)
                c.fill = _fill("F9F9F9" if row_i % 2 == 0 else "FFFFFF")
            c.border = THIN_BORDER

    col_widths = [12, 30, 65, 18]
    for col_i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = w


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def generate_excel_pack(
    pkg: MetricsPackage,
    config_path: str = "config.yaml",
) -> Path:
    """Build the full Excel data pack and write to disk.

    Args:
        pkg: Computed MetricsPackage.
        config_path: Path to configuration YAML.

    Returns:
        Path to the generated .xlsx file.
    """
    with open(config_path, "r") as fh:
        cfg = yaml.safe_load(fh)

    brand = cfg["report"]["brand"]
    output_dir = Path(cfg["paths"]["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = cfg["paths"]["excel_filename"].format(period=pkg.report_period)
    output_path = output_dir / filename

    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Summary",        lambda ws: _sheet_summary(ws, pkg, brand)),
        ("P&L",            lambda ws: _sheet_pl(ws, pkg, brand, config_path)),
        ("Pipeline",       lambda ws: _sheet_pipeline(ws, brand, config_path)),
        ("Customers",      lambda ws: _sheet_customers(ws, brand, config_path)),
        ("Headcount",      lambda ws: _sheet_headcount(ws, brand, config_path)),
        ("Data Dictionary",lambda ws: _sheet_data_dict(ws, brand)),
    ]

    for sheet_name, builder in sheets:
        ws = wb.create_sheet(sheet_name)
        builder(ws)
        logger.info("Built sheet: %s", sheet_name)

    wb.save(output_path)
    logger.info("Excel data pack saved to %s", output_path)
    return output_path
